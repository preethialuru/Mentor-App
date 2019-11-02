import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()
import click
import openpyxl
from onlineapp.models import College, Student,MockTest1

c = College.objects.all()
for e in College.objects.all():
    print(e.acronym+10*" "+e.contact)
cnt = College.objects.count()
print(cnt)
print("college location")
print(College.objects.order_by('-acronym')[:5])

@click.group()
@click.pass_context
def cli(ctx):
   pass

@cli.command('importcollege')
def onlinedb():
    wb1 = openpyxl.load_workbook("students.xlsx")
    sheet1 = wb1.worksheets[2]
    for i in range(2, sheet1.max_row+1):
        college_data = []
        for j in range(1, sheet1.max_column + 1):
            college_data.append(sheet1.cell(row=i, column=j).value)
        c=College(name=college_data[0],location=college_data[2],acronym=college_data[1],contact=college_data[3])
        c.save()

@cli.command('importstudent')
def onlinedb():
    wb1 = openpyxl.load_workbook("students.xlsx")
    sheet1 = wb1.worksheets[0]
    for i in range(2, sheet1.max_row+1):
        students_data = []
        for j in range(1, sheet1.max_column + 1):
            students_data.append(sheet1.cell(row=i, column=j).value)
        s = College.objects.get(acronym=students_data[1])
        c=Student(name=students_data[0],email=students_data[2],db_folder=students_data[3],college=s)
        c.save()

@cli.command('importmocktest')
def onlinedb():
    wb1 = openpyxl.load_workbook("marks.xlsx")
    sheet1 = wb1.worksheets[0]
    #print((sheet1.cell(row=2,column=1).value).split('_')[2])
    for i in range(2,sheet1.max_row+1):
            m = MockTest1()
            m.student = Student.objects.get(db_folder=(((sheet1.cell(row=i,column=1).value).split('_'))[2]))
            m.problem1=sheet1.cell(row=i,column=2).value
            m.problem2 = sheet1.cell(row=i, column=3).value
            m.problem3 = sheet1.cell(row=i, column=4).value
            m.problem4 = sheet1.cell(row=i, column=5).value
            m.total = sheet1.cell(row=i, column=6).value
            m.save()

if __name__ == "__main__":
    cli()















































































































